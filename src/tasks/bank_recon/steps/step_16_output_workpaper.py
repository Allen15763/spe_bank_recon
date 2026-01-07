"""
Step 17: 輸出工作底稿
輸出 Daily Check Excel、Entry Excel，並寫入 Google Sheets
"""

from typing import Dict, Any
from pathlib import Path
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter

from src.core.pipeline import PipelineStep, StepResult, StepStatus
from src.core.pipeline.context import ProcessingContext
from src.core.datasources import GoogleSheetsManager
from src.utils import get_logger, config_manager


class OutputWorkpaperStep(PipelineStep):
    """
    輸出工作底稿步驟
    
    功能:
    1. 輸出 Daily Check Excel
    2. 輸出 Entry Excel
    3. 寫入 Google Sheets（acquiring_charge_raw）
    4. 寫入 Google Sheets（大entry_raw）
    """
    
    def __init__(self, config: Dict[str, Any] = None, **kwargs):
        super().__init__(**kwargs)
        self.config = config or {}
        self.logger = get_logger("OutputWorkpaperStep")
    
    def execute(self, context: ProcessingContext) -> StepResult:
        try:
            self.logger.info("=" * 60)
            self.logger.info("開始輸出工作底稿")
            self.logger.info("=" * 60)
            
            # 取得參數
            output_path = self.config.get('output', {}).get('path', './output/')
            daily_check_filename = context.get_variable('daily_check_filename')
            entry_filename = context.get_variable('entry_filename')
            entry_sheets = context.get_variable('entry_sheets')
            current_month = context.get_variable('current_month')
            
            # 確保輸出目錄存在
            output_dir = Path(output_path)
            output_dir.mkdir(parents=True, exist_ok=True)
            
            output_files = []
            
            # =================================================================
            # 寫入 Google Sheets
            # =================================================================
            gs_config = self.config.get('google_sheets', {})
            
            if gs_config.get('enabled', False):
                try:
                    cred_path = config_manager.get('general', 'cred_path')
                    spreadsheet_url = gs_config.get('spreadsheet_url')
                    
                    gs_manager = GoogleSheetsManager(
                        credentials_path=cred_path,
                        spreadsheet_url=spreadsheet_url
                    )
                    
                    output_config = gs_config.get('output', {})
                    modes_config = gs_config.get('output', {}).get('modes', {})
                    
                    # 3.1 寫入 acquiring_charge_raw (append)
                    acquiring_sheet = output_config.get('acquiring_charge_raw_sheet', 'acquiring_charge_raw')
                    df_apcc_summary_long = context.get_auxiliary_data('apcc_summary_long')
                    
                    if df_apcc_summary_long is not None:
                        is_append = modes_config.get('acquiring_charge_raw', 'append') == 'append'
                        
                        # 準備資料
                        df_to_write = df_apcc_summary_long.copy()
                        df_to_write['period'] = current_month
                        df_to_write['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        
                        # 會計確認無誤後才上傳
                        # gs_manager.write_data(df_to_write, acquiring_sheet, is_append=is_append)
                        self.logger.info(f"已寫入 Google Sheets: {acquiring_sheet}")
                    
                    # 3.2 寫入 大entry_raw (append)
                    big_entry_sheet = output_config.get('big_entry_raw_sheet', '大entry_raw')
                    df_entry_long = context.get_auxiliary_data('entry_long')
                    
                    if df_entry_long is not None:
                        is_append = modes_config.get('big_entry_raw', 'append') == 'append'
                        
                        # 準備資料
                        df_to_write = df_entry_long.copy()
                        df_to_write['period'] = current_month
                        df_to_write['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        
                        # 會計確認無誤後才上傳
                        # gs_manager.write_data(df_to_write, big_entry_sheet, is_append=is_append)
                        self.logger.info(f"已寫入 Google Sheets: {big_entry_sheet}")
                    
                    # 3.3 寫入 acquiring_charge_sum_display (overwrite)
                    display_sheet = output_config.get('acquiring_charge_sum_display_sheet', 
                                                      'acquiring_charge_sum_display')
                    df_summary = context.get_auxiliary_data('df_apcc_summary_fin')
                    
                    if df_summary is not None:
                        is_append = modes_config.get('acquiring_charge_sum_display', 'overwrite') == 'append'
                        gs_manager.write_data(df_summary.astype(object).fillna(''), 
                                              display_sheet, 
                                              is_append=is_append,
                                              clear_range="A1:Q9")

                        self.logger.info(f"已寫入 Google Sheets: {display_sheet}")

                    # 各項指標的樞紐分析表原始資料，上傳雲表
                    self._write_analysis_materials(context, gs_manager)
                    # 輸出到Excel的大entry同步更新至雲表
                    self._write_big_entry(context, gs_manager)
                    
                    # 輸出APCC手續費到雲表； 
                    """注意: 執行一次就好!!! 記得關掉"""
                    # self._write_apcc_data(context, gs_manager)
                    
                except Exception as e:
                    self.logger.warning(f"寫入 Google Sheets 失敗: {e}")
                    context.add_warning(f"Google Sheets 寫入失敗: {e}")

            # =================================================================
            # 寫入 Excel
            # =================================================================
            entry_path = output_dir / entry_filename
            self.logger.info(f"輸出 Entry: {entry_path}")

            self._export_to_excel(
                context, 
                gs_manager, 
                entry_path, 
                output_files, 
                self.logger,
                entry_sheets
            )
            
            # =================================================================
            # 摘要
            # =================================================================
            self.logger.info("\n" + "=" * 60)
            self.logger.info("工作底稿輸出完成")
            self.logger.info(f"  輸出檔案數: {len(output_files)}")
            for f in output_files:
                self.logger.info(f"  - {f}")
            self.logger.info("=" * 60 + "\n")
            
            return StepResult(
                step_name=self.name,
                status=StepStatus.SUCCESS,
                message="工作底稿輸出完成",
                metadata={
                    'output_files': output_files,
                    # 'daily_check_path': str(daily_check_path),
                    'entry_path': str(entry_path),
                    'google_sheets_written': gs_config.get('enabled', True),
                    'output_at': datetime.now().isoformat()
                }
            )
            
        except Exception as e:
            self.logger.error(f"輸出工作底稿失敗: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
            
            return StepResult(
                step_name=self.name,
                status=StepStatus.FAILED,
                error=e,
                message=str(e)
            )
        
    def _write_analysis_materials(self, context, service):
        # Summary的累積分析資訊
        try:
            sheet_name = 'cc_net_rev'
            service.write_data(
                context.get_auxiliary_data('net_cc_rev').astype(object).fillna(''), 
                sheet_name=sheet_name, 
                is_append=False
            )
            self.logger.info(f"已寫入 Google Sheets: {sheet_name}")
        except Exception as e:
            self.logger.error(f"輸出{sheet_name}失敗: {e}")
            import traceback
            self.logger.error(traceback.format_exc())

        try:
            sheet_name = 'spe_proportion'
            service.write_data(
                context.get_auxiliary_data('spe_charge_proportion').astype(object).fillna(''), 
                sheet_name=sheet_name, 
                is_append=False,
                clear_range="A:F"
            )
        except Exception as e:
            self.logger.error(f"輸出{sheet_name}失敗: {e}")
            import traceback
            self.logger.error(traceback.format_exc())

        try:
            sheet_name = 'wp_proportion'
            service.write_data(
                context.get_auxiliary_data('acquiring_proportion').astype(object).fillna(''), 
                sheet_name=sheet_name, 
                is_append=False,
                clear_range="A:E"
            )
        except Exception as e:
            self.logger.error(f"輸出{sheet_name}失敗: {e}")
            import traceback
            self.logger.error(traceback.format_exc())

    def _write_big_entry(self, context, service):
        try:
            sheet_name = '大entry_display'
            service.write_data(
                context.get_auxiliary_data('big_entry').astype(object).fillna(''), 
                sheet_name=sheet_name, 
                is_append=False
            )
        except Exception as e:
            self.logger.error(f"輸出{sheet_name}失敗: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
        
    def _export_to_excel(self, context, gs_manager, entry_path, output_files, logger, entry_sheets):
        """輸出資料到Excel並應用格式"""
        
        # 定義所有需要輸出的工作表配置
        sheets_config = [
            ('big_entry', '大entry'),
            ('dfr_balance_check', '104171 vs DFR'),
            ('category_validation', '分類驗證'),
            ('entry_validation', 'Entry Source Validation', True),  # 進Entry Transformer前的驗證 # True表示需要轉換為DataFrame
            ('entry_temp', 'entry_temp'),
            ('entry_long_temp', 'entry_long_temp'),
            ('dfr_raw', 'dfr'),
            ('dfr_with_balance', 'dfr_detail'),  # DFR 整理明細，含(dfr_result加上)餘額與當日變動、FRR的手續費與匯費
            ('dfr_wp', 'dfr_wp'),
            ('apcc_acquiring_charge', 'APCC手續費'),                      # SPE Charge SPT
            ('apcc_validate_frr', 'apcc_validate_frr'),                  # APCC Validate FRR
            ('apcc_summary', 'Trust account fee回補手續費'),              # Trust account fee回補normal手續費與調扣的版本
            ('trust_account_fee', 'trust_account_fee', False, True),      # 最後的True表示index=True
            ('validation_summary', 'validate_frr_summary'),              # Validate FRR
            ('validate_frr_handling_fee', 'validate_frr_handling_fee'),  # Validate FRR Handling Fee
            ('validate_frr_net_billing', 'validate_frr_net_billing'),    # Validate FRR Net Billing
            ('frr_handling_fee', 'frr_handling_fee'),                    # FRR Handling Fee
            ('frr_remittance_fee', 'frr_remittance_fee'),                # FRR Remittance Fee
            ('frr_long_format', 'extracted_from_frr'),                   # FRR Long Format
            ('acquiring_charge_history', 'acquiring_charge_history'),    # 歷史紀錄參考
            ('apcc_history', 'APCC手續費_history'),                      # 歷史紀錄參考
        ]
        
        with pd.ExcelWriter(entry_path, engine='openpyxl') as writer:
            # 寫入所有工作表
            for config in sheets_config:
                data_key = config[0]
                sheet_name = config[1]
                to_dataframe = config[2] if len(config) > 2 else False
                use_index = config[3] if len(config) > 3 else False
                
                data = context.get_auxiliary_data(data_key)
                
                if data is not None:
                    if to_dataframe:
                        data = pd.DataFrame(data, index=[0])
                    data.to_excel(writer, sheet_name=sheet_name, index=use_index)
            
            # 特殊處理：從Google Sheet讀取summary
            df_summary = gs_manager.get_data('acquiring_charge_sum_display', 'A1:Q11')
            if df_summary is not None:
                df_summary.to_excel(writer, sheet_name='summary_01', index=False)
        
        # 重新開啟檔案以應用格式
        from openpyxl import load_workbook
        workbook = load_workbook(entry_path)
        
        # 對每個工作表應用格式
        for sheet_name in workbook.sheetnames:
            ExcelFormatter.apply_formats(workbook, sheet_name)

        # 重新排序工作表
        desired_order = entry_sheets  # 定義順序
        for idx, sheet_name in enumerate(desired_order):
            if sheet_name in workbook.sheetnames:
                workbook.move_sheet(sheet_name, offset=idx - workbook.index(workbook[sheet_name]))
        
        # 儲存格式化後的檔案
        workbook.save(entry_path)
        workbook.close()
        
        output_files.append(str(entry_path))
        logger.info("Entry Excel 輸出完成（含格式設定）")

    def _write_apcc_data(self, context, service):
        try:
            sheet_name = 'APCC 手續費'
            service.write_data(
                context.get_auxiliary_data('apcc_acquiring_charge_DW').astype(object).fillna(''), 
                sheet_name=sheet_name, 
                is_append=True
            )
        except Exception as e:
            self.logger.error(f"輸出{sheet_name}失敗: {e}")
            import traceback
            self.logger.error(traceback.format_exc())

        try:
            sheet_name = 'APCC 手續費驗證'
            df = context.get_auxiliary_data('apcc_validate_frr')
            cols = ['bank_code', 'subtotal_wp', 'bank_frr', 'subtotal_frr', 'diff']
            df = df[cols].sort_values(by='bank_code').assign(end_date=context.get_variable('end_date'))
            service.write_data(
                df.astype(object).fillna(''), 
                sheet_name=sheet_name, 
                is_append=True
            )
        except Exception as e:
            self.logger.error(f"輸出{sheet_name}失敗: {e}")
            import traceback
            self.logger.error(traceback.format_exc())


class ExcelFormatter:
    """Excel格式設定管理類
    
    如果需要新增格式，需在 SHEET_FORMATS 中新增配置
    '新工作表': {
        'freeze_panes': 'B2',
        'formats': [
            {'columns': ['A'], 'format': ExcelFormatter.DATE_FORMAT},
            {'range': 'B2:F100', 'format': ExcelFormatter.NUMBER_FORMAT_CUSTOM}
        ]
    }
    
    """
    
    # 自定義數字格式
    NUMBER_FORMAT_CUSTOM = '#,##0_);[Red](#,##0)'
    PERCENTAGE_FORMAT = '0.00%'
    DATE_FORMAT = 'yyyy-mm-dd'
    
    # 格式設定配置
    SHEET_FORMATS = {
        '大entry': {
            'freeze_panes': 'D2',
            'formats': [
                {'range': 'D2:ZZ1000', 'format': NUMBER_FORMAT_CUSTOM}
            ]
        },
        '104171 vs DFR': {
            'formats': [
                {'columns': ['B', 'C', 'D', 'E'], 'format': NUMBER_FORMAT_CUSTOM}
            ]
        },
        'entry_temp': {
            'formats': [
                {'columns': 'B:Z', 'format': NUMBER_FORMAT_CUSTOM}
            ]
        },
        'entry_long_temp': {
            'formats': [
                {'columns': ['A'], 'format': DATE_FORMAT},
                {'columns': ['E'], 'format': NUMBER_FORMAT_CUSTOM}
            ]
        },
        'dfr_detail': {
            'formats': [
                {'columns': ['A'], 'format': DATE_FORMAT},
                {'columns': 'B:Z', 'format': NUMBER_FORMAT_CUSTOM}
            ]
        },
        'APCC手續費': {
            'formats': [
                {'columns': ['B', 'C', 'D', 'E', 'F', 'G', 'I', 'J'], 'format': NUMBER_FORMAT_CUSTOM},
                {'columns': ['H'], 'format': PERCENTAGE_FORMAT}
            ]
        },
        'summary_01': {
            'formats': [
                {'range': 'B2:N7', 'format': NUMBER_FORMAT_CUSTOM},
                {'range': 'P2:P7', 'format': NUMBER_FORMAT_CUSTOM},
                {'range': 'B8:Q9', 'format': PERCENTAGE_FORMAT},
                {'range': 'O2:O7', 'format': PERCENTAGE_FORMAT},
                {'range': 'Q2:Q7', 'format': PERCENTAGE_FORMAT}
            ]
        },
        'Trust account fee回補手續費': {
            'formats': [
                {'columns': 'B:Z', 'format': NUMBER_FORMAT_CUSTOM}
            ]
        },
        'trust_account_fee': {
            'formats': [
                {'columns': 'B:Z', 'format': NUMBER_FORMAT_CUSTOM}
            ]
        }
    }
    
    @staticmethod
    def apply_formats(workbook, sheet_name):
        """對指定工作表應用格式"""
        if sheet_name not in ExcelFormatter.SHEET_FORMATS:
            return
        
        ws = workbook[sheet_name]
        config = ExcelFormatter.SHEET_FORMATS[sheet_name]
        
        # 應用凍結窗格
        if 'freeze_panes' in config:
            ws.freeze_panes = config['freeze_panes']
        
        # 應用格式
        for format_config in config.get('formats', []):
            ExcelFormatter._apply_format(ws, format_config)
    
    @staticmethod
    def _apply_format(ws, format_config):
        """應用單個格式配置"""
        number_format = format_config.get('format')
        
        # 處理範圍格式 (例如: 'B2:N7')
        if 'range' in format_config:
            cell_range = format_config['range']
            for row in ws[cell_range]:
                for cell in row:
                    cell.number_format = number_format
        
        # 處理列格式
        elif 'columns' in format_config:
            columns = format_config['columns']
            
            # 處理單個列或列的列表
            if isinstance(columns, str):
                if ':' in columns:  # 例如: 'B:Z'
                    start_col, end_col = columns.split(':')
                    columns = ExcelFormatter._get_column_range(start_col, end_col)
                else:
                    columns = [columns]
            
            # 應用格式到每一列
            for col in columns:
                col_idx = ExcelFormatter._column_to_index(col)
                for row in range(2, ws.max_row + 1):  # 從第2行開始(跳過標題)
                    cell = ws.cell(row=row, column=col_idx)
                    cell.number_format = number_format
    
    @staticmethod
    def _column_to_index(col_letter):
        """將列字母轉換為索引"""
        from openpyxl.utils import column_index_from_string
        return column_index_from_string(col_letter)
    
    @staticmethod
    def _get_column_range(start_col, end_col):
        """取得列範圍"""
        start_idx = ExcelFormatter._column_to_index(start_col)
        end_idx = ExcelFormatter._column_to_index(end_col)
        return [get_column_letter(i) for i in range(start_idx, end_idx + 1)]
